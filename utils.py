import os
from pathlib import Path
import SimpleITK as sitk
import numpy as np
import torch
from torch.utils.data import Dataset
from torchvision import transforms
import openpyxl


def min_max_norm(src_arr):
    max_val = np.max(np.max(src_arr))
    min_val = np.min(np.min(src_arr))

    norm_arr = (src_arr - min_val) / (max_val - min_val + 1e-10)

    return norm_arr


class ImagesDataset_dicom_crop_first_with_window_1channel(Dataset):
    def __init__(self, folder, window, sz):
        super().__init__()
        self.folder = folder
        self.paths = []
        self.window = window
        self.first_crop_sz = sz

        img_to_read = ['.dcm']
        for path in Path(f'{folder}').glob('**/*'):
            _, ext = os.path.splitext(path)
            if ext.lower() in img_to_read:
                self.paths.append(path)

        print(f'{len(self.paths)} dicom images found')
        print("using window size {}".format(self.window))

    def __len__(self):
        return len(self.paths)

    def __getitem__(self, index):
        path = self.paths[index]
        dicom = sitk.ReadImage(path)
        img = np.squeeze(sitk.GetArrayFromImage(dicom))
        img = center_crop_or_pad_array(img, self.first_crop_sz)
        img[img > self.window] = self.window
        img_norm = min_max_norm(img)

        return torch.from_numpy(np.expand_dims(img_norm.astype(np.float32), 0))



class ImagesDataset_dicom_crop_first_1channel_with_window_infer(Dataset):
    def __init__(self, folder, window, image_sz):
        super().__init__()
        self.folder = folder
        self.paths = []
        self.window = window

        img_to_read = ['.dcm']
        for path in Path(f'{folder}').glob('**/*'):
            _, ext = os.path.splitext(path)
            if ext.lower() in img_to_read:
                self.paths.append(path)
        sorted(self.paths)

        # print(f'{len(self.paths)} dicom images found')
        print("{} images found, using window size {}, image size {}".format(len(self.paths), self.window, image_sz))

        self.transform = transforms.Compose([
            transforms.Resize(image_sz),
            # transforms.CenterCrop(256),

        ])

    def __len__(self):
        return len(self.paths)

    def __getitem__(self, index):
        path = self.paths[index]
        dicom = sitk.ReadImage(path)
        img = np.squeeze(sitk.GetArrayFromImage(dicom))

        img = img.astype(np.float32)
        img[img > self.window] = self.window

        img_norm = min_max_norm(img)
        img = np.expand_dims(img_norm, 0)


        return torch.from_numpy(img)


def findAllFile(path, end_name):
    assert end_name[0] == '.' or end_name[0] == "."
    filelist = []
    for root, ds, fs in os.walk(path):
        for f in fs:
            if f.endswith(end_name):
                fullname = os.path.join(root, f)
                filelist.append(fullname)
    return filelist

def writer_excel(data, name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    header = list(data.keys())
    for col in range(len(header)):
        sheet.cell(row=1, column=col + 1, value=header[col])

    for row in range(len(data[header[0]])):
        for col in range(len(header)):
            sheet.cell(row=row + 2, column=col + 1, value=data[header[col]][row])

    workbook.save(name)


def center_crop_or_pad_array(original_array, sz):
    original_rows = len(original_array)
    original_cols = len(original_array[0])


    new_rows = min(sz, original_rows)
    new_cols = min(sz, original_cols)

    start_row = (original_rows - new_rows) // 2 if original_rows > new_rows else 0
    start_col = (original_cols - new_cols) // 2 if original_cols > new_cols else 0

    new_array = [[0 for _ in range(sz)] for _ in range(sz)]

    # for C++ impl, use `for` instead of np slice
    for i in range(new_rows):
        for j in range(new_cols):
            new_array[i][j] = original_array[start_row + i][start_col + j]

    return np.array(new_array)
